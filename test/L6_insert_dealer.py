from openpyxl import load_workbook, Workbook

# โหลดไฟล์ Excel
wb1 = load_workbook("D:\Boatproject\python-project\L6_250k_Test/L6_21k.xlsx")
wb2 = load_workbook("D:\Boatproject\python-project\dataTest/dataTestDup2100row.xlsx")

# เลือกแผ่นงานแรกในแต่ละไฟล์
ws1 = wb1.active
ws2 = wb2.active

# สร้าง Workbook ใหม่สำหรับผลลัพธ์
wb_result = Workbook()
ws_result = wb_result.active

# อ่านข้อมูลจากไฟล์แรกและใส่ลงในผลลัพธ์
for row_index, row in enumerate(ws1.iter_rows(values_only=True), start=1):
    for col_index, value in enumerate(row, start=1):
        ws_result.cell(row=row_index, column=col_index, value=value)

# อ่านข้อมูลจากไฟล์ที่สองและใส่ต่อในคอลัมน์ถัดไป
for row_index, row in enumerate(ws2.iter_rows(values_only=True), start=1):
    ws_result.cell(row=row_index, column=7, value=row[1])

# บันทึกผลลัพธ์ลงในไฟล์ใหม่
wb_result.save("output.xlsx")
print("Files merged and saved as 'output.xlsx'")